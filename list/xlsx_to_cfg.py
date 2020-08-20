from openpyxl import Workbook
from openpyxl import load_workbook
import sys

print(sys.argv)
if len(sys.argv) < 5:
    print("arguments error ! please run : python xlsx_to_cfg.py xlsx_name sheet_name column_index cfg_name")
    sys.exit()

xls_name = sys.argv[1]
sheet_name = sys.argv[2]
column_index = sys.argv[3]
target_name = sys.argv[4]


wb = Workbook()
ws = wb.active

xl_read = load_workbook(xls_name)

res = open(target_name, mode='w')
res.write("instance {"+"\n")


sheet = xl_read.get_sheet_by_name(sheet_name)
max_row = sheet.max_row
for i in range(column_index, max_row+1):
    val = sheet.cell(row=i, column=2).value
    if val is not None:
        res.write(val)
        if i != max_row:
            res.write(",")
        res.write("\n")
res.write("}    {noTiming} ;")
res.close()
