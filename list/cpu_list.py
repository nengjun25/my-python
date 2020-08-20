from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

xlread = load_workbook("synchronizer_list_cpu.xlsx")

res = open("notimingcheck_cpu.cfg", mode='w')
res.write("instance {"+"\n")


tag222 = xlread.get_sheet_by_name("TagV2.2.2")
maxrow = tag222.max_row
for i in range(2, maxrow+1):
    val = tag222.cell(row=i, column=2).value
    if val is not None:
        res.write(val)
        if i != maxrow:
            res.write(",")
        res.write("\n")
res.write("}    {noTiming} ;")
res.close()

