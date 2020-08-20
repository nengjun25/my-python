import sys
import xlwt
import openpyxl

print(sys.argv)
if len(sys.argv) < 5:
    print("arguments error ! please run : python xlsx_to_cfg.py name1.txt name2.xlsx sheet_name column_index")
    sys.exit()


txt_name = sys.argv[1]
xls_name = sys.argv[2]
sheet_name = sys.argv[3]
column_index = sys.argv[4]

# wb = openpyxl.Workbook()
#
# xl_write = openpyxl.load_workbook("synchronizer_list_cpu.xlsx")
# sheet = wb.create_sheet("TagV2.2.2")

workbook = xlwt.Workbook(encoding='gbk')
sheet = workbook.add_sheet(sheet_name)

res = open(txt_name, mode='r')
s = res.read()
s_list = s.splitlines(False)

if len(s_list) > 2:
    del s_list[0]
    s_list.pop()

max_len = 0

font = xlwt.Font()
font.height = 20*11
style = xlwt.XFStyle()
style.font = font

for i in range(0, len(s_list)):
    sheet.write(i, 1, i + 1)
    res = None
    if "," in s_list[i]:
        index = s_list[i].rindex(",")
        res = s_list[i][:index]
    else:
        res = s_list[i]
    sheet.write(i, int(column_index), res, style=style)
    ll = len(s_list[i])
    print(ll)
    if ll > max_len:
        max_len = ll

for j in range(0, i + 5):
    sheet.row(j).height = 20*15
    sheet.col(int(column_index)).width = 256 * (max_len+10)

workbook.save(xls_name)

# sheet = xl_read.get_sheet_by_name(sheet_name)
# max_row = sheet.max_row
# for i in range(column_index, max_row+1):
#     val = sheet.cell(row=i, column=2).value
#     if val is not None:
#         res.write(val)
#         if i != max_row:
#             res.write(",")
#         res.write("\n")
# res.write("}    {noTiming} ;")
# res.close()
