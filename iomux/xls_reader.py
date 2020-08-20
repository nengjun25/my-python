from openpyxl import load_workbook


class XlsReader:
    def __init__(self, file_name, sheet_name):
        self.xls_reader = load_workbook(file_name)
        self.sheet = self.xls_reader.get_sheet_by_name(sheet_name)

    def get_cell_value(self, row, column):
        value = self.sheet.cell(row=row, column=column).value
        if value is not None:
            if type(value) == str:
                value.replace(" ", "")
                if value == "":
                    return None
        return value

    def get_cell_color(self, row, column):
        return self.sheet.cell(row=row, column=column).fill.fgColor.rgb


